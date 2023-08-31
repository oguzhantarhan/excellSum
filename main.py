from openpyxl import Workbook,load_workbook

data=[]
stockData=[]
files=[]
def readpersonnel():
    global files
    wb=load_workbook("PERSONNEL.xlsx")
    ws=wb["personnel"]
    for row in range(1,ws.max_row+1):
        personnel=ws.cell(row,1).value
        files.append(personnel)
def stockSearch(code):
    global stockData
    index=-1
    for product in stockData:
        if code in product:
            index=stockData.index(product)
            break
    return index
def write():
    global data
    wb=Workbook()
    ws=wb.active
    ws.append(["Code","Return count","Outgoing product"])
    for product in data:
        ws.append(product)
        print("Code:{} Return:{} Outgoing:{}".format(product[0],product[1],product[2]))
        
    ws.auto_filter.ref = ws.dimensions
    ws=wb.create_sheet("STOK")
    ws.append(["Code","Stock","Total with return"])
    for product in stockData:
        index=search(product[0])
        if index!=-1:
            returnCount=data[index][1]
            ws.append([product[0],product[1],product[1]+returnCount])
        else:
            ws.append([product[0],product[1],product[1]])
    wb.save("sum.xlsx")

def search(code):
    global data
    index=-1
    for product in data:
        if code in product:
            index=data.index(product)
            break
    return index

def add(code,incoming,outgoing):
    global data
    index=search(code)
    if index==-1:
        data.append([code,incoming,outgoing])
    else:
        data[index][1]=data[index][1]+incoming
        data[index][2]=data[index][2]+outgoing

def stockRead():
    wb=load_workbook("STOCK.xlsx")
    ws=wb.active
    for row in range(2,ws.max_row+1):
        code=ws.cell(row,1).value
        count=ws.cell(row,3).value
        if code==None:
            break
        else:
            global stockData
            index=stockSearch(code)
            if index==-1:
                stockData.append([code,count])
            else:
                stockData[index][1]=stockData[index][1]+count     
def read(files):
    for file in files:
        wb=load_workbook(file)
        ws=wb.sheetnames
        for clinic in ws:
            ws=wb[clinic]
            for row in range(2,ws.max_row+1):
                code=ws.cell(row,1).value
                if code==None:
                    break
                incoming=ws.cell(row,3).value
                outgoing=ws.cell(row,4).value
                print("CODE:{} RETURN:{} OUTGOİNG:{} FİLE:{} SHEET:{}".format(code,incoming,outgoing,file,clinic))
                if code!=None:
                    if incoming==None:
                        incoming=0
                    if outgoing==None:
                        outgoing=0
                    add(code,incoming,outgoing)
readpersonnel()
stockRead()
read(files)
write()
input("press any key  for exit")

